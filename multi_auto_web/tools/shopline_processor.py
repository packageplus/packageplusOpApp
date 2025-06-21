import pandas as pd
import datetime
import os
import openpyxl
import re
import streamlit as st # 引入 Streamlit
from io import BytesIO # 用於處理檔案下載

# 移除 Tkinter 相關的 import
# import tkinter as tk
# from tkinter import filedialog, messagebox

# 您的翻譯字典保持不變
translation_map_shopline = {
    "Taiwan": "台灣",


    # ✅ 北部區域 North Taiwan

    # ✅ 基隆市 Keelung City
    "Keelung City": "基隆市",
    "Ren’ai District": "仁愛區",
    "Xinyi District": "信義區",
    "Zhongzheng District": "中正區",
    "Zhongshan District": "中山區",
    "Anle District": "安樂區",
    "Nuannuan District": "暖暖區",
    "Qidu District": "七堵區",

 # ✅ 新北市 New Taipei City
    "New Taipei City": "新北市",
    "Wanli District": "萬里區",
    "Jinshan District": "金山區",
    "Banqiao District": "板橋區",
    "Xizhi District": "汐止區",
    "Shenkeng District": "深坑區",
    "Shiding District": "石碇區",
    "Ruifang District": "瑞芳區",
    "Pingxi District": "平溪區",
    "Shuangxi District": "雙溪區",
    "Gongliao District": "貢寮區",
    "Xindian District": "新店區",
    "Pinglin District": "坪林區",
    "Wulai District": "烏來區",
    "Yonghe District": "永和區",
    "Zhonghe District": "中和區",
    "Tucheng District": "土城區",
    "Sanxia District": "三峽區",
    "Shulin District": "樹林區",
    "Yingge District": "鶯歌區",
    "Sanchong District": "三重區",
    "Xinzhuang District": "新莊區",
    "Taishan District": "泰山區",
    "Linkou District": "林口區",
    "Luzhou District": "蘆洲區",
    "Wugu District": "五股區",
    "Bali District": "八里區",
    "Tamsui District": "淡水區",
    "Sanzhi District": "三芝區",
    "Shimen District": "石門區",

    # ✅ 台北市 Taipei City
    "Taipei City": "台北市",
    "Zhongzheng District": "中正區",
    "Datong District": "大同區",
    "Zhongshan District": "中山區",
    "Songshan District": "松山區",
    "Daan District": "大安區",
    "Wanhua District": "萬華區",
    "Xinyi District": "信義區",
    "Shilin District": "士林區",
    "Beitou District": "北投區",
    "Neihu District": "內湖區",
    "Nangang District": "南港區",
    "Wenshan District": "文山區",


    # ✅ 桃園市 Taoyuan City
    "Taoyuan City": "桃園市",
    "Zhongli District": "中壢區",
    "Pingzhen District": "平鎮區",
    "Longtan District": "龍潭區",
    "Yangmei District": "楊梅區",
    "Xinwu District": "新屋區",
    "Guanyin District": "觀音區",
    "Taoyuan District": "桃園區",
    "Guishan District": "龜山區",
    "Bade District": "八德區",
    "Daxi District": "大溪區",
    "Fuxing District": "復興區",
    "Dayuan District": "大園區",
    "Luzhu District": "蘆竹區",

    # ✅ 新竹市 Hsinchu City
    "Hsinchu City": "新竹市",
    "East District": "東區",
    "North District": "北區",
    "Xiangshan District": "香山區",

    # ✅ 新竹縣 Hsinchu County
    "Hsinchu County": "新竹縣",
    "Zhubei City": "竹北市",
    "Hukou Township": "湖口鄉",
    "Xinfeng Township": "新豐鄉",
    "Xinpu Township": "新埔鎮",
    "Guanxi Township": "關西鎮",
    "Qionglin Township": "芎林鄉",
    "Baoshan Township": "寶山鄉",
    "Zhudong Township": "竹東鎮",
    "Wufeng Township": "五峰鄉",
    "Hengshan Township": "橫山鄉",
    "Jianshi Township": "尖石鄉",
    "Beipu Township": "北埔鄉",
    "Emei Township": "峨眉鄉",

    # ✅ 宜蘭縣 Yilan County
    "Yilan County": "宜蘭縣",
    "Yilan City": "宜蘭市",
    "Luodong Township": "羅東鎮",
    "Suao Township": "蘇澳鎮",
    "Toucheng Township": "頭城鎮",
    "Jiaoxi Township": "礁溪鄉",
    "Zhuangwei Township": "壯圍鄉",
    "Yuanshan Township": "員山鄉",
    "Dongshan Township": "冬山鄉",
    "Wujie Township": "五結鄉",
    "Sanxing Township": "三星鄉",
    "Datong Township": "大同鄉",
    "Nan'ao Township": "南澳鄉",

    # ✅ 中部區域 Central Taiwan

    # ✅ 台中市 Taichung City
    "Taichung City": "台中市",
    "Central District": "中區",
    "East District": "東區",
    "South District": "南區",
    "West District": "西區",
    "North District": "北區",
    "Beitun District": "北屯區",
    "Xitun District": "西屯區",
    "Nantun District": "南屯區",
    "Taiping District": "太平區",
    "Dali District": "大里區",
    "Wufeng District": "霧峰區",
    "Wuri District": "烏日區",
    "Fengyuan District": "豐原區",
    "Houli District": "后里區",
    "Shigang District": "石岡區",
    "Dongshi District": "東勢區",
    "Heping District": "和平區",
    "Xinshe District": "新社區",
    "Tanzi District": "潭子區",
    "Daya District": "大雅區",
    "Shengang District": "神岡區",
    "Dadu District": "大肚區",
    "Shalu District": "沙鹿區",
    "Longjing District": "龍井區",
    "Wuqi District": "梧棲區",
    "Qingshui District": "清水區",
    "Dajia District": "大甲區",
    "Waipu District": "外埔區",
    "Daan District": "大安區",

    # ✅ 苗栗縣 Miaoli County
    "Miaoli County": "苗栗縣",
    "Miaoli City": "苗栗市",
    "Toufen City": "頭份市",
    "Zhunan Township": "竹南鎮",
    "Houlong Township": "後龍鎮",
    "Zaoqiao Township": "造橋鄉",
    "Touwu Township": "頭屋鄉",
    "Gongguan Township": "公館鄉",
    "Dahu Township": "大湖鄉",
    "Tai'an Township": "泰安鄉",
    "Shitan Township": "獅潭鄉",
    "Nanzhuang Township": "南庄鄉",
    "Sanwan Township": "三灣鄉",
    "Tongluo Township": "銅鑼鄉",
    "Sanyi Township": "三義鄉",
    "Yuanli Township": "苑裡鎮",
    "Tongxiao Township": "通霄鎮",
    "Xihu Township": "西湖鄉",
    "Zhuolan Township": "卓蘭鎮",

    # ✅ 彰化縣 Changhua County
    "Changhua County": "彰化縣",
    "Changhua City": "彰化市",
    "Lugang Township": "鹿港鎮",
    "Hemei Township": "和美鎮",
    "Huatan Township": "花壇鄉",
    "Fenyuan Township": "芬園鄉",
    "Xiushui Township": "秀水鄉",
    "Fuxing Township": "福興鄉",
    "Xianxi Township": "線西鄉",
    "Puyan Township": "埔鹽鄉",
    "Yongjing Township": "永靖鄉",
    "Shetou Township": "社頭鄉",
    "Tianzhong Township": "田中鎮",
    "Beidou Township": "北斗鎮",
    "Tianwei Township": "田尾鄉",
    "Puxin Township": "埔心鄉",
    "Yuanlin City": "員林市",
    "Dacun Township": "大村鄉",
    "Pitou Township": "埤頭鄉",
    "Ershui Township": "二水鄉",
    "Xihu Township": "溪湖鎮",
    "Dacheng Township": "大城鄉",
    "Zhutang Township": "竹塘鄉",
    "Xizhou Township": "溪州鄉",

    # ✅ 南投縣 Nantou County
    "Nantou County": "南投縣",
    "Nantou City": "南投市",
    "Caotun Township": "草屯鎮",
    "Puli Township": "埔里鎮",
    "Zhushan Township": "竹山鎮",
    "Jiji Township": "集集鎮",
    "Lugu Township": "鹿谷鄉",
    "Guoxing Township": "國姓鄉",
    "Ren'ai Township": "仁愛鄉",
    "Xinyi Township": "信義鄉",
    "Yuchi Township": "魚池鄉",
    "Mingjian Township": "名間鄉",
    "Zhongliao Township": "中寮鄉",

    # ✅ 雲林縣 Yunlin County
    "Yunlin County": "雲林縣",
    "Douliu City": "斗六市",
    "Huwei Township": "虎尾鎮",
    "Dounan Township": "斗南鎮",
    "Gukeng Township": "古坑鄉",
    "Citong Township": "莿桐鄉",
    "Siluo Township": "西螺鎮",
    "Xiluo Township": "西螺鎮",
    "Beigang Township": "北港鎮",
    "Shuilin Township": "水林鄉",
    "Baozhong Township": "褒忠鄉",
    "Tuku Township": "土庫鎮",
    "Dapi Township": "大埤鄉",
    "Yuanchang Township": "元長鄉",
    "Mailiao Township": "麥寮鄉",
    "Linnei Township": "林內鄉",
    "Dongshi Township": "東勢鄉",
    "Erlun Township": "二崙鄉",
    "Sihu Township": "四湖鄉",
    "Kouhu Township": "口湖鄉",
    "Taixi Township": "臺西鄉",

    # ✅ 南部區域 Southern Taiwan

    # ✅ 高雄市 Kaohsiung City
    "Kaohsiung City": "高雄市",
    "Xinxing District": "新興區",
    "Qianjin District": "前金區",
    "Lingya District": "苓雅區",
    "Yancheng District": "鹽埕區",
    "Gushan District": "鼓山區",
    "Qianzhen District": "前鎮區",
    "Sanmin District": "三民區",
    "Nanzi District": "楠梓區",
    "Xiaogang District": "小港區",
    "Zuoying District": "左營區",
    "Renwu District": "仁武區",
    "Dashe District": "大社區",
    "Fengshan District": "鳳山區",
    "Daliao District": "大寮區",
    "Linyuan District": "林園區",
    "Niaosong District": "鳥松區",
    "Dashu District": "大樹區",
    "Qishan District": "旗山區",
    "Meinong District": "美濃區",
    "Liugui District": "六龜區",
    "Neimen District": "內門區",
    "Shanlin District": "杉林區",
    "Jiaxian District": "甲仙區",
    "Taoyuan District": "桃源區",
    "Namaxia District": "那瑪夏區",
    "Maolin District": "茂林區",
    "Alian District": "阿蓮區",
    "Tianliao District": "田寮區",
    "Yanchao District": "燕巢區",
    "Qiaotou District": "橋頭區",
    "Ziguan District": "梓官區",
    "Mituo District": "彌陀區",
    "Yongan District": "永安區",
    "Hunei District": "湖內區",
    "Luzhu District": "路竹區",
    "Gangshan District": "岡山區",
    "Xiaogang District": "小港區",

    # ✅ 台南市 Tainan City
    "Tainan City": "台南市",
    "West Central District": "中西區",
    "East District": "東區",
    "South District": "南區",
    "North District": "北區",
    "Anping District": "安平區",
    "Annan District": "安南區",
    "Yongkang District": "永康區",
    "Guiren District": "歸仁區",
    "Xinhua District": "新化區",
    "Zuozhen District": "左鎮區",
    "Yujing District": "玉井區",
    "Nanxi District": "楠西區",
    "Nanhua District": "南化區",
    "Rende District": "仁德區",
    "Guanmiao District": "關廟區",
    "Longqi District": "龍崎區",
    "Guantian District": "官田區",
    "Liujia District": "六甲區",
    "Xigang District": "西港區",
    "Anding District": "安定區",
    "Shanhua District": "善化區",
    "Danei District": "大內區",
    "Shanshang District": "山上區",
    "Xinshi District": "新市區",
    "Sinshih District": "新市區",  # variant spelling
    "Madou District": "麻豆區",
    "Jiali District": "佳里區",
    "Xuejia District": "學甲區",
    "Beimen District": "北門區",
    "Jiangjun District": "將軍區",
    "Qigu District": "七股區",
    "Houbi District": "後壁區",
    "Dongshan District": "東山區",
    "Liuying District": "柳營區",
    "Yanshui District": "鹽水區",
    "Baihe District": "白河區",

    # ✅ 嘉義市 Chiayi City
    "Chiayi City": "嘉義市",
    "East District": "東區",
    "West District": "西區",

    # ✅ 嘉義縣 Chiayi County
    "Chiayi County": "嘉義縣",
    "Taibao City": "太保市",
    "Puzi City": "朴子市",
    "Budai Township": "布袋鎮",
    "Dongshi Township": "東石鄉",
    "Liujiao Township": "六腳鄉",
    "Yizhu Township": "義竹鄉",
    "Lucao Township": "鹿草鄉",
    "Xingang Township": "新港鄉",
    "Minxiong Township": "民雄鄉",
    "Dalin Township": "大林鎮",
    "Meishan Township": "梅山鄉",
    "Zhuqi Township": "竹崎鄉",
    "Fanlu Township": "番路鄉",
    "Alishan Township": "阿里山鄉",
    "Zhongpu Township": "中埔鄉",
    "Shuishang Township": "水上鄉",

    # ✅ 屏東縣 Pingtung County
    "Pingtung County": "屏東縣",
    "Pingtung City": "屏東市",
    "Chaozhou Township": "潮州鎮",
    "Donggang Township": "東港鎮",
    "Hengchun Township": "恆春鎮",
    "Linluo Township": "麟洛鄉",
    "Neipu Township": "內埔鄉",
    "Wanluan Township": "萬巒鄉",
    "Wandan Township": "萬丹鄉",
    "Ligang Township": "里港鄉",
    "Yanpu Township": "鹽埔鄉",
    "Changzhi Township": "長治鄉",
    "Jiuru Township": "九如鄉",
    "Zhutian Township": "竹田鄉",
    "Xinyuan Township": "新園鄉",
    "Fangliao Township": "枋寮鄉",
    "Fangshan Township": "枋山鄉",
    "Checheng Township": "車城鄉",
    "Mudan Township": "牡丹鄉",
    "Shizi Township": "獅子鄉",
    "Manzhou Township": "滿州鄉",
    "Sandimen Township": "三地門鄉",
    "Wutai Township": "霧臺鄉",
    "Majia Township": "瑪家鄉",
    "Taiwu Township": "泰武鄉",
    "Laiyi Township": "來義鄉",
    "Kanding Township": "崁頂鄉",
    "Xinpi Township": "新埤鄉",
    "Jiadong Township": "佳冬鄉",
    "Linbian Township": "林邊鄉",
    "Gaoshu Township": "高樹鄉",
    "Yanpu Township": "鹽埔鄉",
    "Liuqiu Township": "琉球鄉",

    # ✅ 澎湖縣 Penghu County
    "Penghu County": "澎湖縣",
    "Magong City": "馬公市",
    "Huxi Township": "湖西鄉",
    "Baisha Township": "白沙鄉",
    "Xiyu Township": "西嶼鄉",
    "Wangan Township": "望安鄉",
    "Qimei Township": "七美鄉",

    # ✅ 東部區域 Eastern Taiwan

    # ✅ 花蓮縣 Hualien County
    "Hualien County": "花蓮縣",
    "Hualien City": "花蓮市",
    "Ji'an Township": "吉安鄉",
    "Shoufeng Township": "壽豐鄉",
    "Fenglin Township": "鳳林鎮",
    "Guangfu Township": "光復鄉",
    "Fengbin Township": "豐濱鄉",
    "Ruisui Township": "瑞穗鄉",
    "Wanrong Township": "萬榮鄉",
    "Yuli Township": "玉里鎮",
    "Zhuoxi Township": "卓溪鄉",
    "Fuli Township": "富里鄉",
    "Xincheng Township": "新城鄉",

    # ✅ 台東縣 Taitung County
    "Taitung County": "台東縣",
    "Taitung City": "台東市",
    "Chenggong Township": "成功鎮",
    "Guanshan Township": "關山鎮",
    "Beinan Township": "卑南鄉",
    "Luye Township": "鹿野鄉",
    "Yanping Township": "延平鄉",
    "Taimali Township": "太麻里鄉",
    "Jinfeng Township": "金峰鄉",
    "Dawu Township": "大武鄉",
    "Daren Township": "達仁鄉",
    "Chishang Township": "池上鄉",
    "Donghe Township": "東河鄉",
    "Haiduan Township": "海端鄉",
    "Lanyu Township": "蘭嶼鄉",
    "Green Island Township": "綠島鄉",  # not official name, see below

    # ✅ 離島地區 Outlying Islands

    # ✅ 金門縣 Kinmen County
    "Kinmen County": "金門縣",
    "Jincheng Township": "金城鎮",
    "Jinhu Township": "金湖鎮",
    "Jinning Township": "金寧鄉",
    "Jinsha Township": "金沙鎮",
    "Lieyu Township": "烈嶼鄉",
    "Wuqiu Township": "烏坵鄉",

    # ✅ 連江縣 Lienchiang County (馬祖)
    "Lienchiang County": "連江縣",
    "Nangan Township": "南竿鄉",
    "Beigan Township": "北竿鄉",
    "Juguang Township": "莒光鄉",
    "Dongyin Township": "東引鄉",

}

# 處理核心邏輯函式 (原本的 process_excel_logic)
def process_shopline_excel_logic(file_content, translation_map_param):
    """ 處理 Excel 數據的核心邏輯，不包含 GUI 互動 """
    try:
        # ... 您 simple_excel_app.py 中 process_excel_logic 的所有內容 ...
        # 確保在這裡使用了傳入的 translation_map
        st.info("🔄 正在讀取 Excel 檔案...")
        wb = openpyxl.load_workbook(file_content)
        sheet = wb.active
        st.success("✔ Excel 讀取成功！")

        # ... (其他處理步驟，如插入欄位，填寫公式，地址清理) ...

        # ✅ 翻譯 BC 到 BH 欄的行政區名稱 (確保這裡使用 translation_map)
        bc_col_idx = 55
        bh_col_idx = 60
        st.info("🔄 正在翻譯行政區名稱...")
        for row_idx in range(2, sheet.max_row + 1):
            for col_idx in range(bc_col_idx, bh_col_idx + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    sorted_translations = sorted(translation_map.items(), key=lambda item: len(item[0]), reverse=True)
                    for eng, zh in sorted_translations:
                        if eng in cell.value:
                            cell.value = cell.value.replace(eng, zh)
        st.success("✔ 行政區名稱翻譯完成！")

        return wb
    except Exception as e:
        st.error(f"處理失敗：\n{str(e)}")
        return None


# Streamlit 介面函式 (原本的 Streamlit 應用程式介面部分)
def shopline_excel_app():
    st.header("🐦 Shopline 訂單 Excel 處理工具")
    st.markdown("這個工具可以處理 Shopline 訂單 Excel，進行公式插入、地址清理和地區翻譯。")

    uploaded_file = st.file_uploader("請上傳您要處理的 Shopline 訂單 Excel 檔案 (.xlsx)", type=["xlsx"])

    if uploaded_file is not None:
        if st.button("🚀 開始處理 Shopline 訂單"):
            with st.spinner("檔案正在處理中，請稍候..."):
                # 調用核心處理邏輯
                processed_workbook = process_excel_logic_app1(uploaded_file, translation_map_excel_app1) 

            if processed_workbook:
                st.success("✅ 處理完成！您可以下載結果檔案。")
                today_date = datetime.datetime.now().strftime("%m%d")
                output_filename = f"{today_date}_Shopline訂單.xlsx"

                # 下載按鈕邏輯
                from io import BytesIO
                output_buffer = BytesIO()
                processed_workbook.save(output_buffer)
                output_buffer.seek(0)

                st.download_button(
                    label=f"💾 下載 {output_filename}",
                    data=output_buffer,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.error("❗ 處理失敗，請檢查錯誤訊息。")
    st.markdown("---")
    st.markdown("如有任何問題，可能沒有人可以修XD(再看看)")
